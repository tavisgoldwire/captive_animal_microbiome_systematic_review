#!/usr/bin/env python3
"""
CAPTIVE ANIMAL MICROBIOME — AGENT RESULTS INTEGRATION (v2)
============================================================
Merges AI screening agent results into the master Excel workbook.
Handles multiple JSON inputs (Run 1 + Run 2 rerun).

Usage:
  python integrate_agent_results.py
  python integrate_agent_results.py --excel path/to/workbook.xlsx
  python integrate_agent_results.py --json run1.json rerun.json
  python integrate_agent_results.py --excel workbook.xlsx --json run1.json rerun.json

Requires: openpyxl
"""

import argparse
import json
import copy
from datetime import datetime
from pathlib import Path
from collections import Counter

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── STYLES ──
HEADER_FONT = Font(bold=True, size=11, name="Arial")
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT_WHITE = Font(bold=True, size=11, name="Arial", color="FFFFFF")
AGENT_FILL = PatternFill("solid", fgColor="E2EFDA")   # light green for agent rows
RERUN_FILL = PatternFill("solid", fgColor="D6EAF8")   # light blue for rerun rows
NFT_FILL = PatternFill("solid", fgColor="FFF2CC")      # light yellow for needs review
ERROR_FILL = PatternFill("solid", fgColor="FCE4EC")     # light red for errors
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)


def find_files(script_dir):
    """Auto-detect JSON and Excel files in script directory."""
    jsons = []
    excel = None

    json_candidates = ["todo_20260227_112439_raw.json"]
    for name in json_candidates:
        p = script_dir / name
        if p.exists():
            jsons.append(p)

    rerun_dir = script_dir / "outputs" / "rerun"
    if rerun_dir.exists():
        rerun_files = sorted(rerun_dir.glob("rerun_*_raw.json"))
        if rerun_files:
            jsons.append(rerun_files[-1])

    excel_candidates = [
        "SysReviewPapersCaptiveMicrobiome_1.xlsx",
        "SysReviewPapersCaptiveMicrobiome_INTEGRATED.xlsx",
    ]
    for name in excel_candidates:
        p = script_dir / name
        if p.exists():
            excel = p
            break

    return jsons, excel


def load_assignments_lookup(wb):
    ws = wb["Assignments"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers)}
    lookup = {}
    for row in ws.iter_rows(min_row=2):
        vals = [c.value for c in row]
        key = vals[idx["Key"]]
        if not key:
            continue
        lookup[key] = {
            "year": vals[idx.get("Publication Year", 0)],
            "authors": vals[idx.get("Author", 0)],
            "title": vals[idx.get("Title", 0)],
            "journal": vals[idx.get("Journal", 0)],
            "doi": vals[idx.get("DOI", 0)],
            "citations": vals[idx.get("Citations", 0)],
            "abstract": vals[idx.get("Abstract Note", 0)],
        }
    return lookup


def load_existing_dois(wb):
    included_dois = set()
    excluded_dois = set()
    ws = wb["IncludedPapers"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    doi_idx = headers.index("DOI")
    for row in ws.iter_rows(min_row=2):
        d = row[doi_idx].value
        if d:
            included_dois.add(normalize_doi(d))
    ws2 = wb["ExcludedPapers"]
    headers2 = [c.value for c in next(ws2.iter_rows(min_row=1, max_row=1))]
    doi_idx2 = headers2.index("DOI")
    for row in ws2.iter_rows(min_row=2):
        d = row[doi_idx2].value
        if d:
            excluded_dois.add(normalize_doi(d))
    return included_dois, excluded_dois


def normalize_doi(doi):
    if not doi:
        return ""
    doi = str(doi).strip().lower()
    doi = doi.replace("https://doi.org/", "").replace("http://doi.org/", "")
    return doi


def find_last_filled_row(ws, col=3):
    last = 1
    for row in ws.iter_rows(min_row=2, min_col=col, max_col=col):
        if row[0].value is not None:
            last = row[0].row
    return last


def year_to_int(y):
    if y is None:
        return None
    try:
        return int(float(y))
    except (ValueError, TypeError):
        return None


def main():
    parser = argparse.ArgumentParser(description="Integrate agent results into master Excel")
    parser.add_argument("--json", nargs="+", default=None,
                        help="Path(s) to agent result JSON files")
    parser.add_argument("--excel", type=str, default=None,
                        help="Path to master Excel workbook")
    parser.add_argument("--output", type=str, default=None,
                        help="Output Excel path (default: same dir as input)")
    args = parser.parse_args()

    script_dir = Path(__file__).resolve().parent

    if args.json:
        json_paths = [Path(j) for j in args.json]
    else:
        json_paths, _ = find_files(script_dir)

    if args.excel:
        excel_src = Path(args.excel)
    else:
        _, excel_src = find_files(script_dir)

    if not json_paths:
        print("ERROR: No JSON files found. Use --json path/to/results.json")
        return
    if not excel_src or not excel_src.exists():
        print("ERROR: No Excel file found. Use --excel path/to/workbook.xlsx")
        return

    output_path = Path(args.output) if args.output else (
        excel_src.parent / "SysReviewPapersCaptiveMicrobiome_INTEGRATED.xlsx"
    )
    queue_path = excel_src.parent / "human_review_queue.xlsx"

    print(f"Excel source: {excel_src}")
    print(f"JSON files:   {[str(p) for p in json_paths]}")
    print(f"Output:       {output_path}")

    all_records = []
    for jp in json_paths:
        if not jp.exists():
            print(f"  WARNING: {jp} not found, skipping")
            continue
        with open(jp) as f:
            data = json.load(f)
        print(f"  Loaded {len(data)} records from {jp.name}")
        all_records.extend(data)

    # Deduplicate by zotero_key — later records (rerun) take priority
    seen_keys = {}
    for rec in all_records:
        zk = rec.get("zotero_key", "")
        if zk:
            seen_keys[zk] = rec
    deduped = list(seen_keys.values())
    print(f"\nTotal records after dedup: {len(deduped)} (from {len(all_records)} raw)")

    wb = openpyxl.load_workbook(excel_src)
    assignments = load_assignments_lookup(wb)
    inc_dois, exc_dois = load_existing_dois(wb)
    all_existing = inc_dois | exc_dois

    agent_excludes = []
    agent_includes = []
    agent_nft = []
    agent_errors = []

    for rec in deduped:
        zk = rec.get("zotero_key", "")
        meta = assignments.get(zk, {})
        doi = normalize_doi(rec.get("doi") or meta.get("doi", ""))

        if "error" in rec:
            agent_errors.append((rec, meta, doi))
        elif rec.get("screening_decision") == "exclude":
            if doi not in all_existing:
                agent_excludes.append((rec, meta, doi))
        elif rec.get("screening_decision") == "include":
            if doi not in all_existing:
                agent_includes.append((rec, meta, doi))
        elif rec.get("screening_decision") == "needs_full_text":
            if doi not in all_existing:
                agent_nft.append((rec, meta, doi))

    print(f"\nNew records to integrate (excluding already-in-Excel):")
    print(f"  Excludes:        {len(agent_excludes)}")
    print(f"  Includes:        {len(agent_includes)}")
    print(f"  Needs full text: {len(agent_nft)}")
    print(f"  Errors:          {len(agent_errors)}")

    # ═══════════════════════════════════════════════════
    # 1. APPEND EXCLUSIONS
    # ═══════════════════════════════════════════════════
    print("\n[1/4] Appending agent exclusions to ExcludedPapers...")
    ws_exc = wb["ExcludedPapers"]
    exc_last = find_last_filled_row(ws_exc, col=1)

    added_exc = 0
    for rec, meta, doi in agent_excludes:
        exc_last += 1
        confidence = rec.get("screening_confidence", "")
        is_rerun = rec.get("rerun", False)
        label = f"AI-R2-{confidence}" if is_rerun else (f"AI-{confidence}" if confidence else "AI")
        fill = RERUN_FILL if is_rerun else AGENT_FILL

        ws_exc.cell(row=exc_last, column=1, value=year_to_int(meta.get("year")))
        ws_exc.cell(row=exc_last, column=2, value=meta.get("authors", ""))
        ws_exc.cell(row=exc_last, column=3, value=meta.get("title", ""))
        ws_exc.cell(row=exc_last, column=4, value=meta.get("journal", ""))
        ws_exc.cell(row=exc_last, column=5, value="")
        ws_exc.cell(row=exc_last, column=6, value=doi if doi else "")
        ws_exc.cell(row=exc_last, column=7, value=rec.get("screening_reason", ""))
        ws_exc.cell(row=exc_last, column=8, value=label)

        for col in range(1, 9):
            ws_exc.cell(row=exc_last, column=col).fill = fill
            ws_exc.cell(row=exc_last, column=col).font = Font(name="Arial", size=11)
        added_exc += 1

    print(f"  Added {added_exc} agent exclusions")

    # ═══════════════════════════════════════════════════
    # 2. APPEND INCLUDES
    # ═══════════════════════════════════════════════════
    print("\n[2/4] Appending agent includes to IncludedPapers...")
    ws_inc = wb["IncludedPapers"]
    inc_last = find_last_filled_row(ws_inc, col=3)

    FIELD_MAP = {
        "publication_year": 0, "authors": 1, "title": 2, "journal": 3,
        "citations": 4, "doi": 5, "animal_common_name": 6,
        "animal_taxon_scientific": 7, "subspecies": 8, "captive_n": 9,
        "wild_n": 10, "captivity_setting": 11, "geographic_location": 12,
        "mixed_captive_wild": 13, "sample_type": 14, "microbiome_method": 15,
        "interventions": 16, "longitudinal": 17, "healthy_vs_diseased": 18,
        "notes": 19,
    }

    added_inc = 0
    for rec, meta, doi in agent_includes:
        rows = rec.get("rows", [])
        if not rows:
            continue
        is_rerun = rec.get("rerun", False)
        fill = RERUN_FILL if is_rerun else AGENT_FILL
        label = "AI-R2" if is_rerun else "AI-partial"

        for extracted_row in rows:
            inc_last += 1
            for field, col_idx in FIELD_MAP.items():
                val = extracted_row.get(field)
                if val == "NEEDS_FULL_TEXT":
                    cell = ws_inc.cell(row=inc_last, column=col_idx + 1, value="NEEDS_FULL_TEXT")
                    cell.fill = NFT_FILL
                elif isinstance(val, list):
                    ws_inc.cell(row=inc_last, column=col_idx + 1,
                               value=", ".join(str(v) for v in val))
                elif isinstance(val, bool):
                    ws_inc.cell(row=inc_last, column=col_idx + 1,
                               value="Yes" if val else "No")
                elif val is not None:
                    ws_inc.cell(row=inc_last, column=col_idx + 1, value=val)

            if not extracted_row.get("publication_year") and meta.get("year"):
                ws_inc.cell(row=inc_last, column=1, value=year_to_int(meta["year"]))
            if not extracted_row.get("authors") and meta.get("authors"):
                ws_inc.cell(row=inc_last, column=2, value=meta["authors"])
            if not extracted_row.get("doi") and doi:
                ws_inc.cell(row=inc_last, column=6, value=doi)

            ws_inc.cell(row=inc_last, column=21, value=label)

            for col in range(1, 22):
                cell = ws_inc.cell(row=inc_last, column=col)
                if cell.fill == PatternFill():
                    cell.fill = fill
                cell.font = Font(name="Arial", size=11)
            added_inc += 1

    print(f"  Added {added_inc} agent-extracted rows")

    # ═══════════════════════════════════════════════════
    # 3. REVIEW QUEUE
    # ═══════════════════════════════════════════════════
    print("\n[3/4] Creating AgentReviewQueue sheet...")
    if "AgentReviewQueue" in wb.sheetnames:
        del wb["AgentReviewQueue"]
    ws_q = wb.create_sheet("AgentReviewQueue")

    q_headers = [
        "Priority", "Status", "Zotero Key", "DOI", "Title", "Journal",
        "Year", "Authors", "Agent Decision", "Agent Confidence",
        "Agent Reason", "Has Abstract", "Action Required"
    ]
    for col, h in enumerate(q_headers, 1):
        cell = ws_q.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    q_row = 1

    for rec, meta, doi in agent_errors:
        q_row += 1
        has_abs = "Yes" if meta.get("abstract") else "No"
        ws_q.cell(row=q_row, column=1, value=1)
        ws_q.cell(row=q_row, column=2, value="ERROR")
        ws_q.cell(row=q_row, column=3, value=rec.get("zotero_key", ""))
        ws_q.cell(row=q_row, column=4, value=doi)
        ws_q.cell(row=q_row, column=5, value=meta.get("title", ""))
        ws_q.cell(row=q_row, column=6, value=meta.get("journal", ""))
        ws_q.cell(row=q_row, column=7, value=year_to_int(meta.get("year")))
        ws_q.cell(row=q_row, column=8, value=meta.get("authors", ""))
        ws_q.cell(row=q_row, column=9, value="error")
        ws_q.cell(row=q_row, column=10, value="")
        ws_q.cell(row=q_row, column=11, value=str(rec.get("error", ""))[:100])
        ws_q.cell(row=q_row, column=12, value=has_abs)
        ws_q.cell(row=q_row, column=13, value="Human screen — JSON parse errors")
        for col in range(1, 14):
            ws_q.cell(row=q_row, column=col).fill = ERROR_FILL

    nft_medium = [(r, m, d) for r, m, d in agent_nft if r.get("screening_confidence") == "medium"]
    nft_low = [(r, m, d) for r, m, d in agent_nft if r.get("screening_confidence") != "medium"]

    for rec, meta, doi in nft_medium:
        q_row += 1
        has_abs = "Yes" if meta.get("abstract") else "No"
        ws_q.cell(row=q_row, column=1, value=2)
        ws_q.cell(row=q_row, column=2, value="NEEDS_FULL_TEXT")
        ws_q.cell(row=q_row, column=3, value=rec.get("zotero_key", ""))
        ws_q.cell(row=q_row, column=4, value=doi)
        ws_q.cell(row=q_row, column=5, value=meta.get("title", ""))
        ws_q.cell(row=q_row, column=6, value=meta.get("journal", ""))
        ws_q.cell(row=q_row, column=7, value=year_to_int(meta.get("year")))
        ws_q.cell(row=q_row, column=8, value=meta.get("authors", ""))
        ws_q.cell(row=q_row, column=9, value="needs_full_text")
        ws_q.cell(row=q_row, column=10, value="medium")
        ws_q.cell(row=q_row, column=11, value=rec.get("screening_reason", ""))
        ws_q.cell(row=q_row, column=12, value=has_abs)
        ws_q.cell(row=q_row, column=13, value="Human screen — agent unsure")
        for col in range(1, 14):
            ws_q.cell(row=q_row, column=col).fill = NFT_FILL

    for rec, meta, doi in nft_low:
        q_row += 1
        has_abs = "Yes" if meta.get("abstract") else "No"
        ws_q.cell(row=q_row, column=1, value=3)
        ws_q.cell(row=q_row, column=2, value="NEEDS_FULL_TEXT")
        ws_q.cell(row=q_row, column=3, value=rec.get("zotero_key", ""))
        ws_q.cell(row=q_row, column=4, value=doi)
        ws_q.cell(row=q_row, column=5, value=meta.get("title", ""))
        ws_q.cell(row=q_row, column=6, value=meta.get("journal", ""))
        ws_q.cell(row=q_row, column=7, value=year_to_int(meta.get("year")))
        ws_q.cell(row=q_row, column=8, value=meta.get("authors", ""))
        ws_q.cell(row=q_row, column=9, value="needs_full_text")
        ws_q.cell(row=q_row, column=10, value="low")
        ws_q.cell(row=q_row, column=11, value=rec.get("screening_reason", ""))
        ws_q.cell(row=q_row, column=12, value=has_abs)
        ws_q.cell(row=q_row, column=13, value="Human screen — title only, no abstract")
        for col in range(1, 14):
            ws_q.cell(row=q_row, column=col).fill = NFT_FILL

    for col_idx, h in enumerate(q_headers, 1):
        ws_q.column_dimensions[get_column_letter(col_idx)].width = max(len(h) + 4, 15)
    ws_q.column_dimensions["E"].width = 50
    ws_q.column_dimensions["K"].width = 50
    ws_q.column_dimensions["M"].width = 40
    ws_q.auto_filter.ref = f"A1:M{q_row}"

    print(f"  Review queue: {q_row - 1} papers")
    print(f"    Errors:     {len(agent_errors)}")
    print(f"    NFT medium: {len(nft_medium)}")
    print(f"    NFT low:    {len(nft_low)}")

    # ═══════════════════════════════════════════════════
    # 4. SUMMARY SHEET
    # ═══════════════════════════════════════════════════
    print("\n[4/4] Updating PipelineSummary...")
    if "PipelineSummary" in wb.sheetnames:
        del wb["PipelineSummary"]
    ws_s = wb.create_sheet("PipelineSummary", 0)

    total_resolved = len(inc_dois) + len(exc_dois) + added_exc + added_inc
    remaining = len(agent_nft) + len(agent_errors)

    summary_data = [
        ("CAPTIVE ANIMAL MICROBIOME — PIPELINE STATUS", ""),
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("", ""),
        ("CORPUS", ""),
        ("Total papers", 1355),
        ("", ""),
        ("RESOLVED", ""),
        ("Human included", f"{len(inc_dois)} papers"),
        ("Human excluded", f"{len(exc_dois)} papers"),
        ("AI excluded (added)", f"{added_exc} papers"),
        ("AI included (added rows)", f"{added_inc} rows"),
        ("Approx resolved", f"~{total_resolved}"),
        ("Completion", f"{min(total_resolved/1355*100, 100):.0f}%"),
        ("", ""),
        ("REMAINING", ""),
        ("Review queue", f"{remaining} papers"),
        ("  Errors", f"{len(agent_errors)}"),
        ("  NFT medium", f"{len(nft_medium)}"),
        ("  NFT low", f"{len(nft_low)}"),
        ("", ""),
        ("COLOR CODING", ""),
        ("Green rows", "AI Run 1"),
        ("Blue rows", "AI Run 2 (rerun)"),
        ("Yellow cells", "NEEDS_FULL_TEXT"),
        ("Red rows (queue)", "Errors"),
    ]

    for row_idx, (label, value) in enumerate(summary_data, 1):
        cell_a = ws_s.cell(row=row_idx, column=1, value=label)
        cell_b = ws_s.cell(row=row_idx, column=2, value=value)
        cell_a.font = Font(name="Arial", size=11)
        cell_b.font = Font(name="Arial", size=11)
        if label and not value and label == label.upper():
            cell_a.font = Font(name="Arial", size=12, bold=True)

    ws_s.column_dimensions["A"].width = 45
    ws_s.column_dimensions["B"].width = 35

    # ── SAVE ──
    print(f"\nSaving to {output_path}...")
    wb.save(output_path)

    print(f"Creating review queue at {queue_path}...")
    wb_q = openpyxl.Workbook()
    ws_src = wb["AgentReviewQueue"]
    ws_dst = wb_q.active
    ws_dst.title = "ReviewQueue"
    for row in ws_src.iter_rows():
        for cell in row:
            new_cell = ws_dst.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy.copy(cell.font)
                new_cell.fill = copy.copy(cell.fill)
                new_cell.alignment = copy.copy(cell.alignment)
    for col_letter in ws_src.column_dimensions:
        ws_dst.column_dimensions[col_letter].width = ws_src.column_dimensions[col_letter].width
    if ws_src.auto_filter.ref:
        ws_dst.auto_filter.ref = ws_src.auto_filter.ref
    wb_q.save(queue_path)

    print(f"\n{'='*60}")
    print(f"  INTEGRATION COMPLETE")
    print(f"{'='*60}")
    print(f"  Excluded added:  {added_exc}")
    print(f"  Included added:  {added_inc} rows")
    print(f"  Review queue:    {q_row - 1} papers")
    print(f"  Completion:      ~{min(total_resolved/1355*100, 100):.0f}%")
    print(f"\n  {output_path}")
    print(f"  {queue_path}")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()